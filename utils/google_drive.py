# utils/google_drive.py
import json
import mimetypes
import os
import re
import socket
import time
import warnings
import hashlib
import tempfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path, PurePosixPath
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from utils.rutas import (
    RUTA_OAUTH_CONFIG as _OAUTH_CLIENT_CONFIG_PATH,
    RUTA_OAUTH_TOKEN  as _OAUTH_TOKEN_PATH,
    RUTA_STREAMLIT_SECRETS as _STREAMLIT_SECRETS_PATH,
)
from utils.fechas import normalizar_fecha

SCOPES = [
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/spreadsheets",
]
_BASE_DIR              = Path(__file__).resolve().parent.parent
_ENV_OAUTH_CLIENT_JSON = "GOOGLE_OAUTH_CLIENT_JSON"
_ENV_OAUTH_CLIENT_PATH = "GOOGLE_OAUTH_CLIENT_PATH"
_GOOGLE_SHEETS_MIME_TYPE = "application/vnd.google-apps.spreadsheet"
_EXCEL_XLSX_MIME_TYPE    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

_OAUTH_FLOW_CACHE_TTL_SECONDS = 15 * 60
_OAUTH_FLOW_CACHE = {}
_OAUTH_RESULT_CACHE_TTL_SECONDS = 15 * 60
_OAUTH_RESULT_CACHE = {}
_OAUTH_BRIDGE_DIR = Path(tempfile.gettempdir()) / "hojas_verificacion_oauth_bridge"
_OAUTH_TOKEN_STORE_PATH = Path(tempfile.gettempdir()) / "hojas_verificacion_google_drive_token.json"


def _state_key(state):
    estado = str(state or "").strip()
    if not estado:
        return ""
    return hashlib.sha256(estado.encode("utf-8")).hexdigest()


def _ruta_flow_oauth(state):
    clave = _state_key(state)
    if not clave:
        return None
    return _OAUTH_BRIDGE_DIR / f"flow_{clave}.json"


def _ruta_resultado_oauth(state):
    clave = _state_key(state)
    if not clave:
        return None
    return _OAUTH_BRIDGE_DIR / f"result_{clave}.json"


def _guardar_json_seguro(ruta, payload):
    try:
        _OAUTH_BRIDGE_DIR.mkdir(parents=True, exist_ok=True)
        ruta.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
        return True
    except Exception:
        return False


def _cargar_json_seguro(ruta):
    if not ruta or not ruta.exists() or not ruta.is_file():
        return None
    try:
        return json.loads(ruta.read_text(encoding="utf-8"))
    except Exception:
        return None


def _expirado(payload, ttl_segundos):
    try:
        ts = float(payload.get("ts", 0))
    except Exception:
        return True
    return (time.time() - ts) > ttl_segundos


def guardar_token_oauth_local(credenciales_info, usuario=None):
    # Deshabilitado a propósito: la app la usan muchos correos distintos.
    # Persistir el token en un archivo compartido del servidor haría que
    # cualquier usuario que recargue la página heredara la sesión de Drive
    # de otra persona. Cada sesión de navegador debe autenticarse por su cuenta.
    return False


def cargar_token_oauth_local():
    # Deshabilitado: ver nota en guardar_token_oauth_local().
    return None


def limpiar_token_oauth_local():
    # Por seguridad, si existiera un token viejo de una versión anterior
    # del código, lo borra al primer uso; no vuelve a escribir ninguno.
    try:
        if _OAUTH_TOKEN_STORE_PATH.exists():
            _OAUTH_TOKEN_STORE_PATH.unlink()
        if _OAUTH_TOKEN_PATH.exists():
            _OAUTH_TOKEN_PATH.unlink()
        return True
    except Exception:
        return False

def _limpiar_cache_oauth_expirado():
    ahora = time.time()
    expirados = [
        estado
        for estado, payload in _OAUTH_FLOW_CACHE.items()
        if ahora - payload.get("ts", 0) > _OAUTH_FLOW_CACHE_TTL_SECONDS
    ]
    for estado in expirados:
        _OAUTH_FLOW_CACHE.pop(estado, None)


def guardar_flow_config_oauth(state, flow_config):
    estado = str(state or "").strip()
    if not estado or not isinstance(flow_config, dict):
        return

    _limpiar_cache_oauth_expirado()
    payload = {
        "ts": time.time(),
        "flow_config": dict(flow_config),
    }
    _OAUTH_FLOW_CACHE[estado] = payload

    ruta = _ruta_flow_oauth(estado)
    if ruta:
        _guardar_json_seguro(ruta, payload)


def obtener_flow_config_oauth(state):
    estado = str(state or "").strip()
    if not estado:
        return None

    _limpiar_cache_oauth_expirado()
    payload = _OAUTH_FLOW_CACHE.get(estado)
    if payload and not _expirado(payload, _OAUTH_FLOW_CACHE_TTL_SECONDS):
        return dict(payload.get("flow_config") or {})

    ruta = _ruta_flow_oauth(estado)
    payload_archivo = _cargar_json_seguro(ruta)
    if not payload_archivo:
        return None

    if _expirado(payload_archivo, _OAUTH_FLOW_CACHE_TTL_SECONDS):
        try:
            ruta.unlink(missing_ok=True)
        except Exception:
            pass
        return None

    _OAUTH_FLOW_CACHE[estado] = payload_archivo
    return dict(payload_archivo.get("flow_config") or {})


def _limpiar_cache_resultado_oauth_expirado():
    ahora = time.time()
    expirados = [
        estado
        for estado, payload in _OAUTH_RESULT_CACHE.items()
        if ahora - payload.get("ts", 0) > _OAUTH_RESULT_CACHE_TTL_SECONDS
    ]
    for estado in expirados:
        _OAUTH_RESULT_CACHE.pop(estado, None)


def guardar_resultado_oauth(state, credenciales_info, usuario=None, error=None):
    estado = str(state or "").strip()
    if not estado or not isinstance(credenciales_info, dict):
        return

    _limpiar_cache_resultado_oauth_expirado()
    payload = {
        "ts": time.time(),
        "credenciales": dict(credenciales_info),
        "usuario": dict(usuario or {}),
        "error": str(error or "").strip(),
    }
    _OAUTH_RESULT_CACHE[estado] = payload

    ruta = _ruta_resultado_oauth(estado)
    if ruta:
        _guardar_json_seguro(ruta, payload)


def obtener_resultado_oauth(state, consume=True):
    estado = str(state or "").strip()
    if not estado:
        return None

    _limpiar_cache_resultado_oauth_expirado()
    payload = _OAUTH_RESULT_CACHE.get(estado)
    if not payload or _expirado(payload, _OAUTH_RESULT_CACHE_TTL_SECONDS):
        ruta = _ruta_resultado_oauth(estado)
        payload_archivo = _cargar_json_seguro(ruta)
        if not payload_archivo:
            return None

        if _expirado(payload_archivo, _OAUTH_RESULT_CACHE_TTL_SECONDS):
            try:
                ruta.unlink(missing_ok=True)
            except Exception:
                pass
            return None

        payload = payload_archivo
        _OAUTH_RESULT_CACHE[estado] = payload

    if consume:
        _OAUTH_RESULT_CACHE.pop(estado, None)
        ruta = _ruta_resultado_oauth(estado)
        if ruta:
            try:
                ruta.unlink(missing_ok=True)
            except Exception:
                pass

    return {
        "credenciales": dict(payload.get("credenciales") or {}),
        "usuario": dict(payload.get("usuario") or {}),
        "error": str(payload.get("error") or "").strip(),
    }


def obtener_ultimo_resultado_oauth(consume=True):
    _limpiar_cache_resultado_oauth_expirado()

    candidato_payload = None
    candidato_estado = None

    for estado, payload in list(_OAUTH_RESULT_CACHE.items()):
        if _expirado(payload, _OAUTH_RESULT_CACHE_TTL_SECONDS):
            _OAUTH_RESULT_CACHE.pop(estado, None)
            continue
        if candidato_payload is None or float(payload.get("ts", 0)) > float(candidato_payload.get("ts", 0)):
            candidato_payload = payload
            candidato_estado = estado

    try:
        if _OAUTH_BRIDGE_DIR.exists() and _OAUTH_BRIDGE_DIR.is_dir():
            for ruta in _OAUTH_BRIDGE_DIR.glob("result_*.json"):
                payload_archivo = _cargar_json_seguro(ruta)
                if not payload_archivo:
                    continue
                if _expirado(payload_archivo, _OAUTH_RESULT_CACHE_TTL_SECONDS):
                    try:
                        ruta.unlink(missing_ok=True)
                    except Exception:
                        pass
                    continue
                if candidato_payload is None or float(payload_archivo.get("ts", 0)) > float(candidato_payload.get("ts", 0)):
                    candidato_payload = payload_archivo
                    candidato_estado = None
    except Exception:
        pass

    if not candidato_payload:
        return None

    if consume:
        if candidato_estado:
            _OAUTH_RESULT_CACHE.pop(candidato_estado, None)
            ruta_estado = _ruta_resultado_oauth(candidato_estado)
            if ruta_estado:
                try:
                    ruta_estado.unlink(missing_ok=True)
                except Exception:
                    pass
        else:
            ruta_archivo = None
            try:
                if _OAUTH_BRIDGE_DIR.exists() and _OAUTH_BRIDGE_DIR.is_dir():
                    archivos = [
                        r for r in _OAUTH_BRIDGE_DIR.glob("result_*.json")
                        if _cargar_json_seguro(r) == candidato_payload
                    ]
                    if archivos:
                        ruta_archivo = archivos[0]
            except Exception:
                ruta_archivo = None
            if ruta_archivo:
                try:
                    ruta_archivo.unlink(missing_ok=True)
                except Exception:
                    pass

    return {
        "credenciales": dict(candidato_payload.get("credenciales") or {}),
        "usuario": dict(candidato_payload.get("usuario") or {}),
        "error": str(candidato_payload.get("error") or "").strip(),
    }


def guardar_error_oauth(state, mensaje):
    estado = str(state or "").strip()
    if not estado:
        return
    guardar_resultado_oauth(estado, {}, usuario={}, error=mensaje)


def normalizar_fecha(valor, fecha_por_defecto):
    if valor is None:
        return fecha_por_defecto

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    if hasattr(valor, "date") and callable(getattr(valor, "date")):
        try:
            return valor.date()
        except Exception:
            return fecha_por_defecto

    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return fecha_por_defecto
        try:
            return datetime.fromisoformat(texto).date()
        except ValueError:
            return fecha_por_defecto

    return fecha_por_defecto


def resolver_fecha_referencia_drive(fechas_por_concepto=None, fecha_por_defecto=None):
    fecha_base = normalizar_fecha(fecha_por_defecto, datetime.now().date())
    if not isinstance(fechas_por_concepto, dict) or not fechas_por_concepto:
        return fecha_base, False

    fechas = []
    for valor in fechas_por_concepto.values():
        fecha_normalizada = normalizar_fecha(valor, None)
        if fecha_normalizada is not None:
            fechas.append(fecha_normalizada)

    if not fechas:
        return fecha_base, False

    periodos_unicos = sorted({(fecha.year, fecha.month) for fecha in fechas})
    periodo_seleccionado = max(periodos_unicos)
    fecha_resuelta = date(periodo_seleccionado[0], periodo_seleccionado[1], 1)
    return fecha_resuelta, len(periodos_unicos) > 1


def formatear_nombre_carpeta_documentacion(fecha_mantenimiento):
    fecha_normalizada = normalizar_fecha(fecha_mantenimiento, datetime.now().date())
    mes = _MESES[fecha_normalizada.month - 1]
    return f"Documentación MP {mes} {fecha_normalizada.year}"


def construir_ruta_documentacion(fecha_mantenimiento=None):
    fecha_normalizada = normalizar_fecha(fecha_mantenimiento, datetime.now().date())
    mes = _MESES[fecha_normalizada.month - 1]
    return ["Documentación MP", str(fecha_normalizada.year), mes]


def cargar_client_config(desde_json):
    client_config = json.loads(desde_json)
    if not isinstance(client_config, dict) or not any(k in client_config for k in ("installed", "web")):
        raise ValueError("El archivo OAuth no tiene el formato esperado de Google.")
    return client_config


def _obtener_secret_streamlit(clave):
    try:
        import streamlit as st
    except Exception:
        return None

    try:
        if clave in st.secrets:
            return st.secrets[clave]

        if "google_drive" in st.secrets:
            secretos_drive = st.secrets["google_drive"]
            if hasattr(secretos_drive, "get"):
                return secretos_drive.get(clave)
            if clave in secretos_drive:
                return secretos_drive[clave]
    except Exception:
        return None

    return None


def _obtener_secret_desde_archivo(clave):
    if not _STREAMLIT_SECRETS_PATH.exists() or not _STREAMLIT_SECRETS_PATH.is_file():
        return None

    try:
        import tomllib

        secretos = tomllib.loads(_STREAMLIT_SECRETS_PATH.read_text(encoding="utf-8"))
        if clave in secretos:
            return secretos.get(clave)

        secretos_drive = secretos.get("google_drive")
        if isinstance(secretos_drive, dict):
            return secretos_drive.get(clave)
    except Exception:
        return None

    return None


def _resolver_ruta_externa_config(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None
    return Path(os.path.expandvars(os.path.expanduser(texto)))


def _cargar_client_config_desde_archivo(ruta_config):
    if not ruta_config or not ruta_config.exists() or not ruta_config.is_file():
        return None

    with ruta_config.open("r", encoding="utf-8") as archivo:
        return cargar_client_config(archivo.read())


def _resolver_ruta_client_config_drive():
    # Primero usa la ruta canónica y, si no existe, tolera una doble extensión común en Windows.
    candidatos = [_OAUTH_CLIENT_CONFIG_PATH, _OAUTH_CLIENT_CONFIG_PATH.with_suffix(".json.json")]
    for ruta in candidatos:
        if ruta.exists() and ruta.is_file():
            return ruta
    return _OAUTH_CLIENT_CONFIG_PATH


def obtener_ruta_client_config_drive():
    ruta_externa = _resolver_ruta_externa_config(os.getenv(_ENV_OAUTH_CLIENT_PATH))
    if ruta_externa is not None:
        return ruta_externa

    ruta_secreto = _resolver_ruta_externa_config(_obtener_secret_streamlit("google_oauth_client_path"))
    if ruta_secreto is not None:
        return ruta_secreto

    ruta_secreto_archivo = _resolver_ruta_externa_config(_obtener_secret_desde_archivo("google_oauth_client_path"))
    if ruta_secreto_archivo is not None:
        return ruta_secreto_archivo

    return _resolver_ruta_client_config_drive()


def resolver_client_config_drive():
    json_entorno = str(os.getenv(_ENV_OAUTH_CLIENT_JSON, "")).strip()
    if json_entorno:
        return cargar_client_config(json_entorno), f"variable de entorno {_ENV_OAUTH_CLIENT_JSON}"

    json_secreto = _obtener_secret_streamlit("google_oauth_client_json")
    if json_secreto:
        return cargar_client_config(str(json_secreto)), "secreto de Streamlit google_oauth_client_json"

    json_secreto_archivo = _obtener_secret_desde_archivo("google_oauth_client_json")
    if json_secreto_archivo:
        return cargar_client_config(str(json_secreto_archivo)), f"archivo {_STREAMLIT_SECRETS_PATH} (google_oauth_client_json)"

    ruta_externa = _resolver_ruta_externa_config(os.getenv(_ENV_OAUTH_CLIENT_PATH))
    if ruta_externa and ruta_externa.exists() and ruta_externa.is_file():
        return _cargar_client_config_desde_archivo(ruta_externa), f"variable de entorno {_ENV_OAUTH_CLIENT_PATH} ({ruta_externa})"

    ruta_secreto = _resolver_ruta_externa_config(_obtener_secret_streamlit("google_oauth_client_path"))
    if ruta_secreto and ruta_secreto.exists() and ruta_secreto.is_file():
        return _cargar_client_config_desde_archivo(ruta_secreto), f"secreto de Streamlit google_oauth_client_path ({ruta_secreto})"

    ruta_secreto_archivo = _resolver_ruta_externa_config(_obtener_secret_desde_archivo("google_oauth_client_path"))
    if ruta_secreto_archivo and ruta_secreto_archivo.exists() and ruta_secreto_archivo.is_file():
        return _cargar_client_config_desde_archivo(ruta_secreto_archivo), f"archivo {_STREAMLIT_SECRETS_PATH} (google_oauth_client_path -> {ruta_secreto_archivo})"

    ruta_local = _resolver_ruta_client_config_drive()
    if ruta_local.exists() and ruta_local.is_file():
        return _cargar_client_config_desde_archivo(ruta_local), f"archivo local ({ruta_local})"

    return None, "sin configurar"


def cargar_client_config_local():
    client_config, _ = resolver_client_config_drive()
    return client_config


def autorizar_google_drive(client_config):
    """
    Flujo OAuth compatible con Streamlit Cloud.
    Usa redirect URI en lugar de servidor local.
    """
    from google_auth_oauthlib.flow import Flow
    import streamlit as st

    # Detectar si estamos en la nube o en local
    redirect_uri = _detectar_redirect_uri()

    flow = Flow.from_client_config(
        client_config,
        scopes=SCOPES,
        redirect_uri=redirect_uri,
    )

    # Paso 1 — generar URL de autorización
    auth_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )

    # Guardar state y PKCE para verificar e intercambiar después.
    st.session_state["google_oauth_state"] = state
    flow_config = {
        "client_config": client_config,
        "redirect_uri": redirect_uri,
        "code_verifier": getattr(flow, "code_verifier", None),
    }
    st.session_state["google_oauth_flow_config"] = flow_config
    guardar_flow_config_oauth(state, flow_config)

    return auth_url, flow


def _detectar_redirect_uri():
    """Detecta si la app corre en Streamlit Cloud o en local."""
    try:
        import streamlit as st
        # En Streamlit Cloud existe la variable HOSTNAME
        hostname = st.context.headers.get("host", "localhost:8501")
        if "localhost" in hostname or "127.0.0.1" in hostname:
            return "http://localhost:8501"
        return f"https://{hostname}"
    except Exception:
        return "http://localhost:8501"


def intercambiar_codigo_por_credenciales(code, state, flow_config):
    """
    Intercambia el código de autorización por credenciales.
    Se llama después de que Google redirige de vuelta a la app.
    """
    from google_auth_oauthlib.flow import Flow

    flow = Flow.from_client_config(
        flow_config["client_config"],
        scopes=SCOPES,
        redirect_uri=flow_config["redirect_uri"],
        state=state,
    )

    code_verifier = flow_config.get("code_verifier") if isinstance(flow_config, dict) else None
    if code_verifier:
        flow.code_verifier = code_verifier

    flow.fetch_token(code=code)
    credentials = flow.credentials
    return json.loads(credentials.to_json())


def _construir_credenciales_google(credentials_info):
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials

    credentials = Credentials.from_authorized_user_info(credentials_info, SCOPES)
    if credentials.expired and credentials.refresh_token:
        credentials.refresh(Request())
        credentials_info = json.loads(credentials.to_json())

    return credentials, credentials_info


def construir_servicio_drive(credentials_info):
    from googleapiclient.discovery import build
    from google_auth_httplib2 import AuthorizedHttp
    import httplib2

    credentials, credentials_info = _construir_credenciales_google(credentials_info)
    http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=180))
    service = build("drive", "v3", http=http, cache_discovery=False)
    return service, credentials_info


def construir_servicio_sheets(credentials_info):
    from googleapiclient.discovery import build
    from google_auth_httplib2 import AuthorizedHttp
    import httplib2

    credentials, credentials_info = _construir_credenciales_google(credentials_info)
    http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=180))
    service = build("sheets", "v4", http=http, cache_discovery=False)
    return service, credentials_info


def construir_servicios_google(credentials_info):
    from googleapiclient.discovery import build
    from google_auth_httplib2 import AuthorizedHttp
    import httplib2

    credentials, credentials_info = _construir_credenciales_google(credentials_info)
    drive_http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=180))
    sheets_http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=180))
    drive_service = build("drive", "v3", http=drive_http, cache_discovery=False)
    sheets_service = build("sheets", "v4", http=sheets_http, cache_discovery=False)
    return drive_service, sheets_service, credentials_info


def obtener_usuario_conectado(service):
    try:
        info = service.about().get(fields="user(displayName,emailAddress)").execute()
        return info.get("user", {})
    except Exception:
        return {}


def _es_error_temporal_drive(exc):
    if isinstance(exc, (TimeoutError, socket.timeout, ConnectionError)):
        return True
    if isinstance(exc, OSError) and getattr(exc, "winerror", None) == 10060:
        return True

    texto = str(exc).lower()
    return any(patron in texto for patron in ["10060", "timed out", "timeout", "connection reset", "connection aborted"])


def _ejecutar_con_reintentos(request, intentos=4):
    ultimo_error = None
    for intento in range(intentos):
        try:
            return request.execute(num_retries=2)
        except Exception as exc:
            ultimo_error = exc
            if not _es_error_temporal_drive(exc) or intento == intentos - 1:
                raise
            time.sleep(2 + intento * 2)

    raise ultimo_error


def es_error_de_scopes_google(exc):
    texto = str(exc).lower()
    return "insufficient authentication scopes" in texto or "request had insufficient authentication scopes" in texto


def es_error_api_sheets_deshabilitada(exc):
    texto = str(exc).lower()
    return (
        "service_disabled" in texto
        or "google sheets api has not been used" in texto
        or ("sheets.googleapis.com" in texto and "disabled" in texto)
    )


def _escapar_valor_query_drive(valor):
    return str(valor).replace("\\", "\\\\").replace("'", "\\'")


def _buscar_archivos_en_carpeta(service, parent_id, nombre, mime_type=None):
    filtros = [
        "trashed = false",
        f"name = '{_escapar_valor_query_drive(nombre)}'",
        f"'{_escapar_valor_query_drive(parent_id)}' in parents",
    ]
    if mime_type:
        filtros.append(f"mimeType = '{_escapar_valor_query_drive(mime_type)}'")

    consulta = " and ".join(filtros)
    respuesta = _ejecutar_con_reintentos(service.files().list(
        q=consulta,
        spaces="drive",
        fields="files(id, name, mimeType, webViewLink)",
        pageSize=100
    ))
    return respuesta.get("files", [])


def obtener_o_crear_carpeta(service, nombre, parent_id=None):
    parent_busqueda = parent_id or "root"
    existentes = _buscar_archivos_en_carpeta(
        service,
        parent_busqueda,
        nombre,
        mime_type="application/vnd.google-apps.folder"
    )
    if existentes:
        return existentes[0]

    metadata = {
        "name": nombre,
        "mimeType": "application/vnd.google-apps.folder",
    }
    metadata["parents"] = [parent_busqueda]

    return _ejecutar_con_reintentos(service.files().create(body=metadata, fields="id, name, webViewLink"))


def obtener_o_crear_ruta_carpetas(service, nombres_carpeta, parent_id=None):
    carpetas = []
    carpeta_padre = parent_id
    for nombre in nombres_carpeta:
        carpeta = obtener_o_crear_carpeta(service, nombre, parent_id=carpeta_padre)
        carpetas.append(carpeta)
        carpeta_padre = carpeta["id"]
    return carpetas


def _separar_nombre_extension(nombre_archivo):
    base, extension = os.path.splitext(nombre_archivo)
    return base, extension


def resolver_nombre_unico_drive(service, nombre_archivo, folder_id, mime_type=None):
    if not _buscar_archivos_en_carpeta(service, folder_id, nombre_archivo, mime_type=mime_type):
        return nombre_archivo

    base, extension = _separar_nombre_extension(nombre_archivo)
    indice = 1
    while True:
        candidato = f"{base} ({indice}){extension}"
        if not _buscar_archivos_en_carpeta(service, folder_id, candidato, mime_type=mime_type):
            return candidato
        indice += 1


def crear_carpeta_unica(service, nombre_base, parent_id=None):
    nombre_final = nombre_base
    if parent_id:
        nombre_final = resolver_nombre_unico_drive(
            service,
            nombre_base,
            parent_id,
            mime_type="application/vnd.google-apps.folder"
        )

    metadata = {
        "name": nombre_final,
        "mimeType": "application/vnd.google-apps.folder",
    }
    if parent_id:
        metadata["parents"] = [parent_id]

    return _ejecutar_con_reintentos(service.files().create(body=metadata, fields="id, name, webViewLink"))


def subir_archivo_bytes(service, nombre_archivo, contenido, folder_id=None, mime_type="application/zip", usar_nombre_unico=True):
    from googleapiclient.http import MediaIoBaseUpload

    nombre_final = nombre_archivo
    if folder_id and usar_nombre_unico:
        nombre_final = resolver_nombre_unico_drive(service, nombre_archivo, folder_id, mime_type=mime_type)

    metadata = {
        "name": nombre_final,
    }
    if folder_id:
        metadata["parents"] = [folder_id]

    media = MediaIoBaseUpload(BytesIO(contenido), mimetype=mime_type, resumable=False)
    archivo = _ejecutar_con_reintentos(service.files().create(
        body=metadata,
        media_body=media,
        fields="id, name, webViewLink"
    ))
    return archivo


def _es_validacion_checkbox_excel(validacion):
    formula1 = str(getattr(validacion, "formula1", "") or "")
    if getattr(validacion, "type", "") != "list":
        return False

    formula_normalizada = formula1.lower().strip()
    tiene_marcas = any(marca in formula1 for marca in ("✓", "✔", "✖", "✘"))
    tiene_na_texto = any(na in formula_normalizada for na in ("n/a", "na", "n a"))
    if tiene_marcas or tiene_na_texto:
        return True

    if formula_normalizada.startswith('"') and formula_normalizada.endswith('"'):
        opciones = [op.strip().lower() for op in formula_normalizada[1:-1].split(",") if op.strip()]
        opciones_set = set(opciones)
        tiene_na = bool(opciones_set & {"n/a", "na", "n a"})
        tiene_check = bool(opciones_set & {"✔", "✓", "✘", "✖"})
        if tiene_na or tiene_check:
            return True

    return False


def _normalizar_formula_validacion(formula):
    texto = str(formula or "").strip()
    if texto.startswith("="):
        texto = texto[1:]
    return texto.strip()


def _normalizar_valor_fecha_google(valor):
    """Convierte fórmulas/valores de fecha de Excel a formato estable para Google Sheets."""
    texto = _normalizar_formula_validacion(valor)
    if not texto:
        return ""

    # Excel suele exportar límites de validación como DATE(yyyy,m,d).
    match = re.fullmatch(r"date\((\d{4}),(\d{1,2}),(\d{1,2})\)", texto, flags=re.IGNORECASE)
    if match:
        y, m, d = map(int, match.groups())
        try:
            return date(y, m, d).strftime("%Y-%m-%d")
        except ValueError:
            return ""

    # Quitar comillas externas si existen.
    if len(texto) >= 2 and texto[0] == '"' and texto[-1] == '"':
        texto = texto[1:-1].strip()

    # Formatos comunes que Sheets sí interpreta como fecha literal.
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", texto):
        return texto
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", texto):
        return texto

    # Si llega otra fórmula (ej. TODAY()), no forzarla en DATE_BETWEEN.
    return ""


def _construir_regla_validacion_google(validacion):
    tipo = str(getattr(validacion, "type", "") or "").strip().lower()
    operador = str(getattr(validacion, "operator", "") or "").strip().lower()
    formula1 = _normalizar_formula_validacion(getattr(validacion, "formula1", ""))
    formula2 = _normalizar_formula_validacion(getattr(validacion, "formula2", ""))

    if tipo == "list":
        if formula1.startswith('"') and formula1.endswith('"'):
            opciones = [op.strip() for op in formula1[1:-1].split(",") if op.strip()]
            if not opciones:
                return None
            return {
                "condition": {
                    "type": "ONE_OF_LIST",
                    "values": [{"userEnteredValue": op} for op in opciones],
                },
                "strict": False,
                "showCustomUi": True,
            }

        if formula1:
            return {
                "condition": {
                    "type": "ONE_OF_RANGE",
                    "values": [{"userEnteredValue": formula1}],
                },
                "strict": False,
                "showCustomUi": True,
            }
        return None

    if tipo == "date":
        mapa_operadores = {
            "between": "DATE_BETWEEN",
            "notbetween": "DATE_NOT_BETWEEN",
            "equal": "DATE_EQ",
            "notequal": "DATE_NOT_EQ",
            "greaterthan": "DATE_AFTER",
            "lessthan": "DATE_BEFORE",
            "greaterthanorequal": "DATE_ON_OR_AFTER",
            "lessthanorequal": "DATE_ON_OR_BEFORE",
        }
        tipo_condicion = mapa_operadores.get(operador, "DATE_IS_VALID")
        fecha1 = _normalizar_valor_fecha_google(formula1)
        fecha2 = _normalizar_valor_fecha_google(formula2)
        values = []

        if fecha1:
            values.append({"userEnteredValue": fecha1})
        if fecha2:
            values.append({"userEnteredValue": fecha2})

        # Si no hay límites de fecha válidos, usar validación genérica de fecha.
        if tipo_condicion in {"DATE_BETWEEN", "DATE_NOT_BETWEEN"} and len(values) < 2:
            tipo_condicion = "DATE_IS_VALID"
            values = []
        elif tipo_condicion in {"DATE_EQ", "DATE_NOT_EQ", "DATE_AFTER", "DATE_BEFORE", "DATE_ON_OR_AFTER", "DATE_ON_OR_BEFORE"} and len(values) < 1:
            tipo_condicion = "DATE_IS_VALID"
            values = []

        regla = {
            "condition": {
                "type": tipo_condicion,
            },
            "strict": False,
            "showCustomUi": True,
        }
        if values:
            regla["condition"]["values"] = values
        return regla

    return None


def _celda_valor_triestado(valor):
    if valor is None:
        return ""

    if isinstance(valor, bool):
        return "✔" if valor else ""

    texto = str(valor).strip()
    if not texto:
        return ""

    texto_norm = texto.lower()
    if texto_norm in {"n/a", "na", "n a"}:
        return "N/A"
    if texto in {"✔", "✓", "✘", "✖"}:
        return "✔" if texto in {"✔", "✓"} else "✘"
    if texto_norm in {"true", "verdadero", "si", "sí", "1", "x"}:
        return "✔"
    if texto_norm in {"false", "falso", "no", "0"}:
        return "✘"

    return texto


def _extraer_configuracion_checkbox_desde_excel(contenido_excel):
    configuracion_por_hoja = {}

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            category=UserWarning,
        )
        workbook = load_workbook(BytesIO(contenido_excel))

    try:
        for hoja in workbook.worksheets:
            data_validations = getattr(hoja, "data_validations", None)
            if data_validations is None:
                continue

            rangos_checkbox = []
            for validacion in list(data_validations.dataValidation):
                if not _es_validacion_checkbox_excel(validacion):
                    continue

                for rango_a1 in str(validacion.sqref).split():
                    min_col, min_row, max_col, max_row = range_boundaries(rango_a1)
                    filas = []
                    for fila in range(min_row, max_row + 1):
                        valores = []
                        for columna in range(min_col, max_col + 1):
                            triestado = _celda_valor_triestado(hoja.cell(row=fila, column=columna).value)
                            if triestado:
                                valores.append({"userEnteredValue": {"stringValue": triestado}})
                            else:
                                valores.append({})
                        filas.append({"values": valores})

                    rangos_checkbox.append({
                        "a1_range": rango_a1,
                        "rows": filas,
                    })

            if rangos_checkbox:
                configuracion_por_hoja[hoja.title] = rangos_checkbox
    finally:
        workbook.close()

    return configuracion_por_hoja


def _extraer_validaciones_generales_desde_excel(contenido_excel):
    validaciones_por_hoja = {}

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            category=UserWarning,
        )
        workbook = load_workbook(BytesIO(contenido_excel))

    try:
        for hoja in workbook.worksheets:
            data_validations = getattr(hoja, "data_validations", None)
            if data_validations is None:
                continue

            reglas = []
            for validacion in list(data_validations.dataValidation):
                if _es_validacion_checkbox_excel(validacion):
                    continue

                regla_google = _construir_regla_validacion_google(validacion)
                if not regla_google:
                    continue

                for rango_a1 in str(validacion.sqref).split():
                    reglas.append({
                        "a1_range": rango_a1,
                        "rule": regla_google,
                    })

            if reglas:
                validaciones_por_hoja[hoja.title] = reglas
    finally:
        workbook.close()

    return validaciones_por_hoja


def _extraer_columnas_checkbox_desde_excel(contenido_excel):
    columnas_por_hoja = {}
    encabezados_objetivo = {"cumple", "no cumple", "no aplica", "n/a", "na", "n a", "si", "no"}
    marcas_checkbox = {"✔", "✓", "✘", "✖", "n/a", "na", "n a"}

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            category=UserWarning,
        )
        workbook = load_workbook(BytesIO(contenido_excel), data_only=False)

    try:
        for hoja in workbook.worksheets:
            max_col = min(hoja.max_column, 120)
            max_row = min(hoja.max_row, 240)
            columnas = {}

            for fila in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    valor = hoja.cell(row=fila, column=col).value
                    if valor is None:
                        continue

                    texto = str(valor).strip()
                    if not texto:
                        continue

                    texto_norm = re.sub(r"[^a-z0-9/]+", " ", texto.lower()).strip()
                    es_encabezado = (
                        texto_norm in encabezados_objetivo
                        or texto_norm.startswith("cumple")
                        or texto_norm.startswith("no cumple")
                        or texto_norm.startswith("no aplica")
                    )
                    es_marca = texto in {"✔", "✓", "✘", "✖"} or texto_norm in marcas_checkbox
                    if not (es_encabezado or es_marca):
                        continue

                    fila_inicio = fila + 1 if es_encabezado else fila
                    if col not in columnas or fila_inicio < columnas[col]:
                        columnas[col] = fila_inicio

            if columnas:
                columnas_por_hoja[hoja.title] = [
                    {
                        "column_index_0": col - 1,
                        "start_row_index_0": max(0, fila_inicio - 1),
                    }
                    for col, fila_inicio in sorted(columnas.items())
                ]
    finally:
        workbook.close()

    return columnas_por_hoja


def _obtener_ids_hojas_google(service_sheets, spreadsheet_id):
    respuesta = _ejecutar_con_reintentos(service_sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties(sheetId,title)"
    ))
    return {
        hoja["properties"]["title"]: hoja["properties"]["sheetId"]
        for hoja in respuesta.get("sheets", [])
        if "properties" in hoja and "title" in hoja["properties"] and "sheetId" in hoja["properties"]
    }


def _construir_grid_range(sheet_id, a1_range):
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    return {
        "sheetId": sheet_id,
        "startRowIndex": min_row - 1,
        "endRowIndex": max_row,
        "startColumnIndex": min_col - 1,
        "endColumnIndex": max_col,
    }


def _normalizar_texto_check_google(valor):
    texto = str(valor or "").strip().lower()
    texto = re.sub(r"[^a-z0-9/]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _detectar_columnas_verificacion_google(data_hoja):
    columnas = {}
    encabezados_objetivo = {"cumple", "no cumple", "no aplica", "n/a", "na", "n a"}

    for fila_idx, fila in enumerate(data_hoja.get("rowData", [])):
        for col_idx, celda in enumerate(fila.get("values", [])):
            texto = _normalizar_texto_check_google(celda.get("formattedValue", ""))
            if (
                texto in encabezados_objetivo
                or texto.startswith("cumple")
                or texto.startswith("no cumple")
                or texto.startswith("no aplica")
            ):
                if col_idx not in columnas or fila_idx < columnas[col_idx]:
                    columnas[col_idx] = fila_idx

    return columnas


def _aplicar_checkboxes_en_google_sheet(service_sheets, spreadsheet_id, configuracion_por_hoja):
    ids_por_hoja = _obtener_ids_hojas_google(service_sheets, spreadsheet_id)
    solicitudes = []

    for titulo_hoja, rangos in configuracion_por_hoja.items():
        sheet_id = ids_por_hoja.get(titulo_hoja)
        if sheet_id is None:
            raise ValueError(f"No se encontró la hoja '{titulo_hoja}' después de convertir el archivo a Google Sheets.")

        for rango in rangos:
            grid_range = _construir_grid_range(sheet_id, rango["a1_range"])
            solicitudes.append({
                "updateCells": {
                    "range": grid_range,
                    "rows": rango["rows"],
                    "fields": "userEnteredValue",
                }
            })
            solicitudes.append({
                "setDataValidation": {
                    "range": grid_range,
                    "rule": {
                        "condition": {
                            "type": "ONE_OF_LIST",
                            "values": [
                                {"userEnteredValue": "✔"},
                                {"userEnteredValue": "✘"},
                                {"userEnteredValue": "N/A"},
                            ],
                        },
                        "strict": False,
                        "showCustomUi": True,
                    }
                }
            })

    if not solicitudes:
        return

    _ejecutar_con_reintentos(service_sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": solicitudes}
    ))


def _aplicar_validaciones_generales_en_google_sheet(service_sheets, spreadsheet_id, validaciones_por_hoja):
    ids_por_hoja = _obtener_ids_hojas_google(service_sheets, spreadsheet_id)
    solicitudes = []

    for titulo_hoja, reglas in validaciones_por_hoja.items():
        sheet_id = ids_por_hoja.get(titulo_hoja)
        if sheet_id is None:
            continue

        for regla in reglas:
            grid_range = _construir_grid_range(sheet_id, regla["a1_range"])
            solicitudes.append({
                "setDataValidation": {
                    "range": grid_range,
                    "rule": regla["rule"],
                }
            })

    if solicitudes:
        _ejecutar_con_reintentos(service_sheets.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": solicitudes}
        ))


def _aplicar_checkboxes_fallback_por_encabezado(service_sheets, spreadsheet_id):
    metadata = _ejecutar_con_reintentos(service_sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)))"
    ))

    solicitudes = []
    for hoja in metadata.get("sheets", []):
        props = hoja.get("properties", {})
        sheet_id = props.get("sheetId")
        titulo = props.get("title", "")
        total_rows = int(props.get("gridProperties", {}).get("rowCount", 1000) or 1000)
        if sheet_id is None or not titulo:
            continue

        muestra = _ejecutar_con_reintentos(service_sheets.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            ranges=[f"'{titulo}'!A1:CB240"],
            includeGridData=True,
            fields="sheets(data(rowData(values(formattedValue))))"
        ))
        sheets_data = muestra.get("sheets", [])
        if not sheets_data:
            continue

        data_hoja = sheets_data[0].get("data", [{}])[0]
        columnas = _detectar_columnas_verificacion_google(data_hoja)
        for col_idx, header_row_idx in columnas.items():
            start_row = header_row_idx + 1
            if start_row >= total_rows:
                continue

            grid_range = {
                "sheetId": sheet_id,
                "startRowIndex": start_row,
                "endRowIndex": total_rows,
                "startColumnIndex": col_idx,
                "endColumnIndex": col_idx + 1,
            }

            solicitudes.append({
                "repeatCell": {
                    "range": grid_range,
                    "cell": {
                        "dataValidation": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": "✔"},
                                    {"userEnteredValue": "✘"},
                                    {"userEnteredValue": "N/A"},
                                ],
                            },
                            "strict": False,
                            "showCustomUi": True,
                        },
                    },
                    "fields": "dataValidation",
                }
            })

    if solicitudes:
        _ejecutar_con_reintentos(service_sheets.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": solicitudes}
        ))


def _aplicar_checkboxes_por_columna_desde_excel(service_sheets, spreadsheet_id, columnas_por_hoja):
    if not columnas_por_hoja:
        return

    metadata = _ejecutar_con_reintentos(service_sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title,gridProperties(rowCount)))"
    ))

    props_por_titulo = {}
    for hoja in metadata.get("sheets", []):
        props = hoja.get("properties", {})
        titulo = props.get("title")
        sheet_id = props.get("sheetId")
        total_rows = int(props.get("gridProperties", {}).get("rowCount", 0) or 0)
        if titulo and sheet_id is not None and total_rows > 0:
            props_por_titulo[titulo] = (sheet_id, total_rows)

    solicitudes = []
    for titulo_hoja, columnas in columnas_por_hoja.items():
        if titulo_hoja not in props_por_titulo:
            continue

        sheet_id, total_rows = props_por_titulo[titulo_hoja]
        for item in columnas:
            col_idx = int(item.get("column_index_0", -1))
            start_row_idx = int(item.get("start_row_index_0", 0))
            if col_idx < 0 or start_row_idx >= total_rows:
                continue

            solicitudes.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_row_idx,
                        "endRowIndex": total_rows,
                        "startColumnIndex": col_idx,
                        "endColumnIndex": col_idx + 1,
                    },
                    "cell": {
                        "dataValidation": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": "✔"},
                                    {"userEnteredValue": "✘"},
                                    {"userEnteredValue": "N/A"},
                                ],
                            },
                            "strict": False,
                            "showCustomUi": True,
                        },
                    },
                    "fields": "dataValidation",
                }
            })

    if solicitudes:
        _ejecutar_con_reintentos(service_sheets.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": solicitudes}
        ))


def subir_excel_como_google_sheet(service_drive, service_sheets, nombre_archivo, contenido, folder_id=None, usar_nombre_unico=True):
    from googleapiclient.http import MediaIoBaseUpload

    nombre_base, _ = _separar_nombre_extension(nombre_archivo)
    nombre_destino = nombre_base or nombre_archivo
    if folder_id and usar_nombre_unico:
        nombre_destino = resolver_nombre_unico_drive(
            service_drive,
            nombre_destino,
            folder_id,
            mime_type=_GOOGLE_SHEETS_MIME_TYPE,
        )

    metadata = {
        "name": nombre_destino,
        "mimeType": _GOOGLE_SHEETS_MIME_TYPE,
    }
    if folder_id:
        metadata["parents"] = [folder_id]

    media = MediaIoBaseUpload(BytesIO(contenido), mimetype=_EXCEL_XLSX_MIME_TYPE, resumable=False)
    archivo = _ejecutar_con_reintentos(service_drive.files().create(
        body=metadata,
        media_body=media,
        fields="id, name, mimeType, webViewLink"
    ))

    try:
        configuracion_por_hoja = _extraer_configuracion_checkbox_desde_excel(contenido)
        validaciones_generales = _extraer_validaciones_generales_desde_excel(contenido)

        if validaciones_generales:
            _aplicar_validaciones_generales_en_google_sheet(service_sheets, archivo["id"], validaciones_generales)
        if configuracion_por_hoja:
            _aplicar_checkboxes_en_google_sheet(service_sheets, archivo["id"], configuracion_por_hoja)
    except Exception:
        try:
            _ejecutar_con_reintentos(service_drive.files().delete(fileId=archivo["id"]))
        except Exception:
            pass
        raise

    return archivo


def _iterar_entradas_zip_validas(contenido_zip):
    with zipfile.ZipFile(BytesIO(contenido_zip), "r") as archivo_zip:
        info_list = [info for info in archivo_zip.infolist() if info.filename and not info.filename.startswith("__MACOSX/")]
        rutas_validas = []
        for info in info_list:
            ruta = PurePosixPath(info.filename)
            partes = [parte for parte in ruta.parts if parte not in ("", ".")]
            if not partes or any(parte == ".." for parte in partes):
                continue
            rutas_validas.append((info, partes))

        prefijo_comun = None
        if rutas_validas:
            primeras_partes = {partes[0] for _, partes in rutas_validas if partes}
            if len(primeras_partes) == 1:
                prefijo_comun = next(iter(primeras_partes))

        for info, partes in rutas_validas:
            yield archivo_zip, info, partes, prefijo_comun


def _es_pdf_bytes(contenido):
    return bytes(contenido[:5]).startswith(b"%PDF-")


def _es_excel_xlsx_bytes(contenido):
    try:
        with zipfile.ZipFile(BytesIO(contenido), "r") as paquete_excel:
            nombres = set(paquete_excel.namelist())
            return "[Content_Types].xml" in nombres and "xl/workbook.xml" in nombres
    except Exception:
        return False


def subir_zip_como_documentos(service, contenido_zip, folder_id, service_sheets=None):
    carpetas_cache = {(): folder_id}
    archivos_subidos = []
    sheets_disponible = service_sheets

    for archivo_zip, info, partes, prefijo_comun in _iterar_entradas_zip_validas(contenido_zip):
            partes_subida = list(partes)
            partes_originales = list(partes)

            # No crear carpeta contenedora del paquete en Drive.
            if prefijo_comun and partes_subida and partes_subida[0] == prefijo_comun:
                partes_subida = partes_subida[1:]
            elif partes_subida and re.match(r"^reporte[_\- ].*", str(partes_subida[0]), flags=re.IGNORECASE):
                partes_subida = partes_subida[1:]

            if not partes_subida:
                if info.is_dir():
                    # Es la carpeta contenedora del paquete; se omite intencionalmente.
                    continue
                # Es un archivo cuyo nombre completo coincidía con el "prefijo" detectado
                # (por ejemplo, un único archivo suelto en la raíz del zip, o un archivo
                # cuyo nombre empieza con "reporte_"). No se debe perder el archivo:
                # se conserva su nombre original sin recortar.
                partes_subida = partes_originales

            if info.is_dir():
                carpeta_actual = folder_id
                ruta_acumulada = []
                for parte in partes_subida:
                    ruta_acumulada.append(parte)
                    clave = tuple(ruta_acumulada)
                    if clave in carpetas_cache:
                        carpeta_actual = carpetas_cache[clave]
                        continue
                    carpeta = obtener_o_crear_carpeta(service, parte, parent_id=carpeta_actual)
                    carpeta_actual = carpeta["id"]
                    carpetas_cache[clave] = carpeta_actual
                continue

            carpeta_actual = folder_id
            ruta_acumulada = []
            for parte in partes_subida[:-1]:
                ruta_acumulada.append(parte)
                clave = tuple(ruta_acumulada)
                if clave not in carpetas_cache:
                    carpeta = obtener_o_crear_carpeta(service, parte, parent_id=carpeta_actual)
                    carpetas_cache[clave] = carpeta["id"]
                carpeta_actual = carpetas_cache[clave]

            nombre_archivo = str(partes_subida[-1]).strip()
            nombre_archivo_lower = nombre_archivo.lower()
            contenido = archivo_zip.read(info.filename)
            es_pdf = nombre_archivo_lower.endswith(".pdf") or _es_pdf_bytes(contenido)
            es_excel = nombre_archivo_lower.endswith(".xlsx") or _es_excel_xlsx_bytes(contenido)

            if not es_pdf and not es_excel:
                continue

            if es_pdf and not nombre_archivo_lower.endswith(".pdf"):
                nombre_archivo = f"{nombre_archivo}.pdf"
                nombre_archivo_lower = nombre_archivo.lower()
            if es_excel and not nombre_archivo_lower.endswith(".xlsx"):
                nombre_archivo = f"{nombre_archivo}.xlsx"
                nombre_archivo_lower = nombre_archivo.lower()

            mime_type = "application/pdf" if es_pdf else _EXCEL_XLSX_MIME_TYPE

            if sheets_disponible and mime_type == _EXCEL_XLSX_MIME_TYPE:
                try:
                    archivo = subir_excel_como_google_sheet(
                        service,
                        sheets_disponible,
                        nombre_archivo,
                        contenido,
                        folder_id=carpeta_actual,
                        usar_nombre_unico=True
                    )
                except Exception as exc:
                    # Fallback automático: si Sheets API no está habilitada o faltan scopes,
                    # sube el archivo .xlsx normal para no bloquear toda la carga.
                    if es_error_de_scopes_google(exc) or es_error_api_sheets_deshabilitada(exc):
                        sheets_disponible = None
                        archivo = subir_archivo_bytes(
                            service,
                            nombre_archivo,
                            contenido,
                            folder_id=carpeta_actual,
                            mime_type=mime_type,
                            usar_nombre_unico=True
                        )
                        archivo["_checkbox_fallback"] = True
                    else:
                        raise
            else:
                archivo = subir_archivo_bytes(
                    service,
                    nombre_archivo,
                    contenido,
                    folder_id=carpeta_actual,
                    mime_type=mime_type,
                    usar_nombre_unico=True
                )
            archivos_subidos.append(archivo)

    return archivos_subidos

# utils/gestor_plantillas.py
import json
import os
import re
import unicodedata
import zipfile
from copy import copy
from io import BytesIO
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as pdf_canvas

from utils.rutas import RUTA_MAPEO, RUTA_PLANTILLAS, RUTA_REPORTES
from utils.fechas import (
    normalizar_fecha,
    formato_mes_anio,
    calcular_siguiente_mantenimiento,
    resolver_fecha_base_por_equipo,
)


# ── Mapeo de plantillas ──────────────────────────────────────────────────────

def cargar_mapeo():
    try:
        with open(RUTA_MAPEO, "r", encoding="utf-8") as archivo:
            return json.load(archivo)
    except Exception as e:
        print(f"Error al cargar mapeo: {e}")
        return None


def _normalizar_concepto(concepto):
    texto = str(concepto or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def obtener_pestana(concepto):
    mapeo = cargar_mapeo()
    if mapeo is None:
        return None

    if concepto in mapeo:
        return mapeo[concepto]

    concepto_norm = _normalizar_concepto(concepto)
    if not concepto_norm:
        return None

    mapeo_normalizado = {}
    for clave, pestana in mapeo.items():
        clave_norm = _normalizar_concepto(clave)
        if clave_norm and clave_norm not in mapeo_normalizado:
            mapeo_normalizado[clave_norm] = pestana

    if concepto_norm in mapeo_normalizado:
        return mapeo_normalizado[concepto_norm]

    for clave_norm, pestana in mapeo_normalizado.items():
        if concepto_norm in clave_norm or clave_norm in concepto_norm:
            return pestana

    return None


# ── Utilidades de nombre de archivo ─────────────────────────────────────────

def _normalizar_nombre_archivo(nombre):
    nombre = str(nombre).strip()
    nombre = unicodedata.normalize("NFKD", nombre)
    nombre = "".join(c for c in nombre if not unicodedata.combining(c))
    nombre = nombre.replace('"', "'")
    nombre = re.sub(r"[\/\\:\*\?<>|]", " ", nombre)
    nombre = re.sub(r"\s+", " ", nombre)
    return nombre.strip()


def _obtener_identificador_activo(equipo):
    for clave in ["# ACTIVO", "ID TINC", "id tinc", "ID", "id"]:
        if clave in equipo and equipo.get(clave):
            return str(equipo.get(clave)).strip()
    return "SIN_ACTIVO"


def _obtener_tipo_activo(equipo):
    return str(equipo.get("CONCEPTO", "SIN_TIPO_ACTIVO")).strip()


def _construir_nombre_hoja(equipo, usados):
    activo   = _normalizar_nombre_archivo(_obtener_identificador_activo(equipo))
    concepto = _normalizar_nombre_archivo(_obtener_tipo_activo(equipo))
    base = f"{activo} {concepto}".strip() or "HOJA"
    base = re.sub(r"[\[\]\*\?/\\:]", " ", base)
    base = re.sub(r"\s+", " ", base).strip()[:31].rstrip() or "HOJA"

    candidato = base
    contador  = 2
    while candidato in usados:
        sufijo    = f"_{contador}"
        candidato = f"{base[:max(1, 31 - len(sufijo))].rstrip()}{sufijo}"
        contador += 1

    usados.add(candidato)
    return candidato


# ── Utilidades de celdas ─────────────────────────────────────────────────────

def _obtener_celda_para_escribir(ws, fila, columna):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= fila <= merged_range.max_row and
            merged_range.min_col <= columna <= merged_range.max_col
        ):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return ws.cell(row=fila, column=columna)


def _escribir_valor(ws, fila, columna, valor):
    _obtener_celda_para_escribir(ws, fila, columna).value = valor


def _quitar_negritas_celda(celda):
    try:
        if celda.font and celda.font.bold:
            nueva_fuente        = copy(celda.font)
            nueva_fuente.bold   = False
            celda.font          = nueva_fuente
    except Exception:
        pass


def _obtener_columna_derecha_etiqueta(ws, fila, columna):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= fila <= merged_range.max_row and
            merged_range.min_col <= columna <= merged_range.max_col
        ):
            return merged_range.max_col + 1
    return columna + 1


def _escribir_valor_con_link(ws, celda_ref, valor, url=None):
    celda = ws[celda_ref]
    if url:
        valor = str(valor or "").replace('"', '""')
        url   = str(url).strip().replace('"', '""')
        celda.value     = f'=HYPERLINK("{url}","{valor}")'
        celda.hyperlink = url
        celda.style     = "Hyperlink"
    else:
        celda.value     = valor
        celda.hyperlink = None


# ── Búsqueda genérica en celdas ──────────────────────────────────────────────

def _normalize_text(s):
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def _buscar_etiqueta_y_escribir(ws, texto_buscado, valor, escribir_arriba=False, usar_celda_derecha=False, url=None):
    """
    Función genérica que recorre la hoja buscando una etiqueta de texto
    y escribe el valor en la celda adyacente (arriba o a la derecha).
    Retorna True si encontró y escribió al menos una vez.
    """
    escrito = False
    try:
        for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in fila:
                if not isinstance(cell.value, str):
                    continue
                norm = _normalize_text(cell.value)
                if texto_buscado not in norm:
                    continue

                if escribir_arriba and cell.row > 1:
                    _escribir_valor(ws, cell.row - 1, cell.column, valor or "")
                    escrito = True
                elif usar_celda_derecha:
                    col_destino   = _obtener_columna_derecha_etiqueta(ws, cell.row, cell.column)
                    celda_destino = _obtener_celda_para_escribir(ws, cell.row, col_destino)
                    if url:
                        _escribir_valor_con_link(ws, celda_destino.coordinate, str(valor or ""), url)
                    else:
                        celda_destino.value = valor
                    escrito = True
    except Exception:
        pass
    return escrito


def _buscar_y_escribir_firmas_por_texto(ws, ingeniero, jefe):
    """Escribe ingeniero y jefe en las celdas encima de sus respectivas etiquetas."""
    try:
        for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in fila:
                if not isinstance(cell.value, str):
                    continue
                norm = _normalize_text(cell.value)
                if "nombre" in norm and "ingeniero" in norm:
                    if cell.row > 1:
                        _escribir_valor(ws, cell.row - 1, cell.column, ingeniero or "")
                elif "nombre" in norm and ("jefe" in norm or "jefe de" in norm or "jefe servicio" in norm):
                    if cell.row > 1:
                        _escribir_valor(ws, cell.row - 1, cell.column, jefe or "")
    except Exception:
        pass


def _buscar_y_escribir_folio_por_texto(ws, folio, url=None):
    """Escribe el folio TiNC en la celda a la derecha de la etiqueta FOLIO."""
    folio_texto = str(folio or "").strip()
    if not folio_texto:
        return False

    escrito = False
    try:
        for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in fila:
                if not isinstance(cell.value, str):
                    continue
                texto = _normalize_text(cell.value)
                texto = re.sub(r"[^a-z0-9]+", " ", texto).strip()
                if texto in {"folio", "no folio", "numero de folio"} or texto.startswith("folio "):
                    col_destino   = _obtener_columna_derecha_etiqueta(ws, cell.row, cell.column)
                    celda_destino = _obtener_celda_para_escribir(ws, cell.row, col_destino)
                    _escribir_valor_con_link(ws, celda_destino.coordinate, folio_texto, url)
                    escrito = True
    except Exception:
        pass
    return escrito


# ── Analizadores ─────────────────────────────────────────────────────────────

def _resolver_serie_desde_fila(fila):
    """Extrae el número de serie desde un diccionario de equipo o analizador."""
    for clave in (
        "serie", "sn", "ns", "n_s", "num_serie", "numero_serie", "numero de serie",
        "SERIE", "SN", "NS", "NUMERO DE SERIE"
    ):
        valor = fila.get(clave, "") if hasattr(fila, "get") else ""
        texto = str(valor or "").strip()
        if texto:
            return texto
    return ""


def _encontrar_fila_encabezado_analizadores(ws):
    encabezados_buscados = {"equipo", "marca", "modelo", "serie", "no. de serie", "numero de serie"}

    def _norm(valor):
        texto = str(valor or "").strip().lower()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(c for c in texto if not unicodedata.combining(c))
        return re.sub(r"\s+", " ", texto)

    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=35):
        encontrados = set()
        for cell in fila:
            if isinstance(cell.value, str):
                texto = _norm(cell.value)
                if any(k in texto for k in ("equipo", "marca", "modelo", "serie")):
                    encontrados.add(texto)
        if len(encontrados) >= 3:
            return fila[0].row

    if ws.max_row >= 44:
        return 44
    if ws.max_row >= 10:
        return 10
    return 1


def _llenar_analizadores(ws, analizadores):
    def _parse_value(value):
        if isinstance(value, dict):
            serie = _resolver_serie_desde_fila(value)
            return {
                "tipo":   value.get("tipo", value.get("Analizador", value.get("analizador", ""))),
                "marca":  value.get("marca", ""),
                "modelo": value.get("modelo", ""),
                "serie":  serie,
            }
        if not value:
            return {"tipo": "", "marca": "", "modelo": "", "serie": ""}
        partes = [p.strip() for p in str(value).split("|")]
        return {
            "tipo":   partes[0] if len(partes) > 0 else "",
            "marca":  partes[1] if len(partes) > 1 else "",
            "modelo": partes[2] if len(partes) > 2 else "",
            "serie":  partes[3] if len(partes) > 3 else "",
        }

    if analizadores is None:
        analizadores = []
    elif isinstance(analizadores, (str, dict)):
        analizadores = [analizadores]

    encabezado   = _encontrar_fila_encabezado_analizadores(ws)
    filas_inicio = encabezado + 1
    columnas     = [1, 2, 12, 21, 30]
    celdas_con_datos = set()

    for idx, valor in enumerate(analizadores):
        fila_actual  = filas_inicio + idx
        analizador   = _parse_value(valor)
        valores_col  = {
            1:  "--",
            2:  analizador.get("tipo", ""),
            12: analizador.get("marca", ""),
            21: analizador.get("modelo", ""),
            30: str(analizador.get("serie", "") or ""),
        }
        for col in columnas:
            celda = _obtener_celda_para_escribir(ws, fila_actual, col)
            celdas_con_datos.add(celda.coordinate)
            celda.value = valores_col.get(col, "")
        _quitar_negritas_celda(_obtener_celda_para_escribir(ws, fila_actual, 30))

    for idx in range(len(analizadores), max(len(analizadores), 6)):
        fila_actual = filas_inicio + idx
        for col in columnas:
            celda = _obtener_celda_para_escribir(ws, fila_actual, col)
            if celda.coordinate not in celdas_con_datos:
                celda.value = None


# ── Imágenes ─────────────────────────────────────────────────────────────────

def _copiar_imagenes_hoja(ws_origen, ws_destino):
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from io import BytesIO
    from PIL import Image as PILImage

    for imagen in getattr(ws_origen, "_images", []) or []:
        try:
            # Leer los bytes de la imagen antes de que el archivo se cierre
            if hasattr(imagen, "ref") and imagen.ref is not None:
                pil_img = PILImage.open(imagen.ref)
                buffer = BytesIO()
                pil_img.save(buffer, format=pil_img.format or "PNG")
                buffer.seek(0)
                nueva_imagen = OpenpyxlImage(buffer)
            else:
                continue

            if hasattr(imagen, "anchor"):
                try:
                    nueva_imagen.anchor = copy(imagen.anchor)
                except Exception:
                    nueva_imagen.anchor = imagen.anchor

            ws_destino.add_image(nueva_imagen)
        except Exception:
            continue


# ── Validaciones ─────────────────────────────────────────────────────────────

def _preparar_fechas_por_texto(ws):
    try:
        validacion_fecha = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
        )
        validacion_fecha.showInputMessage  = False
        validacion_fecha.showErrorMessage  = False
        ws.add_data_validation(validacion_fecha)

        for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in fila:
                if not isinstance(cell.value, str):
                    continue
                texto = _normalize_text(cell.value)
                texto = re.sub(r"[^a-z0-9]+", " ", texto).strip()
                if "fecha" not in texto:
                    continue
                col_destino   = _obtener_columna_derecha_etiqueta(ws, cell.row, cell.column)
                if col_destino > ws.max_column + 1:
                    continue
                celda_destino = _obtener_celda_para_escribir(ws, cell.row, col_destino)
                celda_destino.number_format = "DD/MMM/YYYY"
                validacion_fecha.add(celda_destino)
    except Exception:
        pass


def _aplicar_validacion_verificacion(ws):
    try:
        validacion_estado = DataValidation(
            type="list",
            formula1='"✔,✘,N/A"',
            allow_blank=True,
        )
        validacion_estado.showInputMessage = False
        validacion_estado.showErrorMessage = False
        ws.add_data_validation(validacion_estado)

        encabezados_objetivo = {"cumple", "no cumple", "no aplica", "na", "n a", "n/a", "si", "no"}
        columnas_objetivo    = set()
        max_col = min(ws.max_column, 80)
        max_row = min(ws.max_row, 120)

        for fila in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in fila:
                if not isinstance(cell.value, str):
                    continue
                texto = _normalize_text(cell.value)
                texto = re.sub(r"[^a-z0-9/]+", " ", texto).strip()
                if texto in encabezados_objetivo or texto.startswith(("cumple", "no aplica")):
                    columnas_objetivo.add(cell.column)

        for col in sorted(columnas_objetivo):
            for row in range(1, ws.max_row + 1):
                celda = _obtener_celda_para_escribir(ws, row, col)
                if isinstance(celda.value, str):
                    texto = _normalize_text(celda.value)
                    if texto in encabezados_objetivo or texto.startswith(("cumple", "no aplica")):
                        continue
                if celda.value in (None, "", "✔", "✘", "N/A"):
                    validacion_estado.add(celda)
    except Exception:
        pass


# ── Mapeo estático de celdas por plantilla ────────────────────────────────────

def _obtener_celdas_firma_por_pestana(pestana):
    celdas = {
        "UNIDAD DE ELECTROCIRUGIA":  ("B58",  "Q58"),
        "CENTRIFUGA":                ("B53",  "Q53"),
        "BASCULA":                   ("B47",  "Q47"),
        "CAMA":                      ("B51",  "Q51"),
        "MESA DE EXPLORACION":       ("B56",  "Q56"),
        "SIERRA":                    ("B47",  "Q47"),
        "MAQUINA DE ANESTESIA":      ("B106", "Q106"),
        "DESFIBRILADOR":             ("B79",  "Q79"),
        "MONITOR DE SIGNOS VITALES": ("B56",  "Q56"),
    }
    return celdas.get(str(pestana).strip().upper(), (None, None))


def _obtener_celda_folio_por_pestana(pestana):
    celdas = {
        "UNIDAD DE ELECTROCIRUGIA":  "AF8",
        "CENTRIFUGA":                "AF8",
        "BASCULA":                   "AF8",
        "CAMA":                      "AF8",
        "MESA DE EXPLORACION":       "AF8",
        "SIERRA":                    "AF8",
        "MAQUINA DE ANESTESIA":      "AE8",
        "DESFIBRILADOR":             "AE8",
        "MONITOR DE SIGNOS VITALES": "AE8",
    }
    return celdas.get(str(pestana).strip().upper())


def _obtener_celdas_folio_refuerzo(celda_principal):
    if celda_principal == "AE8":
        return ["AF8"]
    if celda_principal == "AF8":
        return ["AE8"]
    return []


# ── Ubicación ─────────────────────────────────────────────────────────────────

def _construir_ubicacion_sububicacion(equipo):
    ubicacion    = str(equipo.get("UBICACIÓN", "") or "").strip()
    sububicacion = ""
    for clave in ("SUB UBICACIÓN", "SUBUBICACIÓN", "SUB UBICACION", "SUBUBICACION"):
        if clave in equipo and str(equipo.get(clave, "") or "").strip():
            sububicacion = str(equipo.get(clave, "") or "").strip()
            break
    if ubicacion and sububicacion:
        return f"{ubicacion} - {sububicacion}"
    return ubicacion or sububicacion


# ── Llenado de hoja ──────────────────────────────────────────────────────────

def _llenar_hoja_verificacion(ws, equipo, ingeniero, jefe, hospital, nombre_plantilla, analizadores=None):
    ws["G12"] = hospital
    ws["G13"] = _construir_ubicacion_sububicacion(equipo)
    ws["G14"] = equipo.get("# ACTIVO", equipo.get("ID TINC", equipo.get("id tinc", "")))

    folio_tinc = equipo.get("FOLIO TINC", equipo.get("FOLIO", ""))
    url_tinc   = equipo.get("URL TINC",   equipo.get("URL",   ""))

    if str(folio_tinc or "").strip():
        _buscar_y_escribir_folio_por_texto(ws, folio_tinc, url_tinc)
        celda_folio = _obtener_celda_folio_por_pestana(nombre_plantilla)
        if celda_folio:
            _escribir_valor_con_link(ws, celda_folio, folio_tinc, url_tinc)
            for celda_extra in _obtener_celdas_folio_refuerzo(celda_folio):
                try:
                    if not ws[celda_extra].value:
                        _escribir_valor_con_link(ws, celda_extra, folio_tinc, url_tinc)
                except Exception:
                    pass

    ws["AA12"] = equipo.get("MARCA", "")
    ws["AA13"] = equipo.get("MODELO", "")
    ws["AA14"] = equipo.get("No. DE SERIE", "")
    _quitar_negritas_celda(ws["AA14"])

    _buscar_y_escribir_firmas_por_texto(ws, ingeniero, jefe)

    ing_cell, jefe_cell = _obtener_celdas_firma_por_pestana(nombre_plantilla)
    if ing_cell:
        try:
            if not ws[ing_cell].value:
                ws[ing_cell] = ingeniero or ""
        except Exception:
            pass
    if jefe_cell:
        try:
            if not ws[jefe_cell].value:
                ws[jefe_cell] = jefe or ""
        except Exception:
            pass

    if analizadores:
        _llenar_analizadores(ws, analizadores)

    _preparar_fechas_por_texto(ws)
    _aplicar_validacion_verificacion(ws)


# ── Generación de reportes ────────────────────────────────────────────────────

def generar_reporte(equipo, ingeniero, jefe, hospital, analizadores=None):
    try:
        concepto = equipo.get("CONCEPTO", "")
        pestana  = obtener_pestana(concepto)

        if pestana is None:
            return None, f"No hay plantilla asignada para '{concepto}'"

        wb = load_workbook(RUTA_PLANTILLAS)
        if pestana not in wb.sheetnames:
            return None, f"La pestaña '{pestana}' no existe en el archivo de plantillas"

        ws = wb[pestana]
        _llenar_hoja_verificacion(ws, equipo, ingeniero, jefe, hospital,
                                  nombre_plantilla=pestana, analizadores=analizadores)

        os.makedirs(RUTA_REPORTES, exist_ok=True)

        tipo_activo  = _normalizar_nombre_archivo(_obtener_tipo_activo(equipo))
        identificador = _normalizar_nombre_archivo(_obtener_identificador_activo(equipo))
        nombre_archivo = f"{identificador} {tipo_activo}.xlsx"
        ruta_destino   = os.path.join(RUTA_REPORTES, nombre_archivo)

        for hoja in [h for h in wb.sheetnames if h != pestana]:
            del wb[hoja]

        wb.save(ruta_destino)
        return ruta_destino, None
    except Exception as e:
        return None, f"Error al generar reporte: {e}"


def crear_paquete_reporte(equipos, nombre_carpeta, ingeniero, jefe=None, hospital=None,
                          progress_bar=None, status_text=None,
                          hacer_hojas=True, hacer_etiquetas=True,
                          analizador=None,
                          analizadores_por_concepto=None,
                          analizadores_seleccionados=None,
                          fecha_mantenimiento_base=None):

    buffer_zip = BytesIO()
    errores    = []
    exitos     = 0

    if analizadores_por_concepto is None:
        analizadores_por_concepto = {}
    if analizadores_seleccionados is None:
        analizadores_seleccionados = []

    nombre_carpeta = _normalizar_nombre_archivo(nombre_carpeta) or "paquete"

    with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr(f"{nombre_carpeta}/", "")

        if hacer_hojas:
            try:
                wb_base       = load_workbook(RUTA_PLANTILLAS)
                pestañas_base = set(wb_base.sheetnames)
                wb_base.close()
            except Exception as e:
                status_text.empty()
                return buffer_zip, [f"No se pudo abrir la plantilla base: {e}"], 0

            equipos_por_concepto = {}
            for idx, (_, row) in enumerate(equipos.iterrows()):
                activo   = row.get("# ACTIVO", "SIN_ACTIVO")
                concepto = row.get("CONCEPTO", "SIN_CONCEPTO")
                status_text.text(f"Procesando: {activo}")
                pestana = obtener_pestana(concepto)

                if pestana is None:
                    errores.append(f"El equipo {activo} ({concepto}) no tiene plantilla.")
                    progress_bar.progress((idx + 1) / len(equipos))
                    continue
                if pestana not in pestañas_base:
                    errores.append(f"El equipo {activo} ({concepto}) apunta a '{pestana}' que no existe.")
                    progress_bar.progress((idx + 1) / len(equipos))
                    continue

                equipos_por_concepto.setdefault(concepto, {"pestana": pestana, "equipos": []})
                equipos_por_concepto[concepto]["equipos"].append(row.to_dict())
                progress_bar.progress((idx + 1) / len(equipos))

            for concepto, data in equipos_por_concepto.items():
                pestana      = data["pestana"]
                lista_equipos = data["equipos"]
                if not lista_equipos:
                    continue

                tipo_equipo    = _normalizar_nombre_archivo(str(concepto or "SIN_TIPO_ACTIVO"))
                carpeta_equipo = f"{nombre_carpeta}/{tipo_equipo}"
                zip_file.writestr(f"{carpeta_equipo}/", "")

                try:
                    wb_concepto = load_workbook(RUTA_PLANTILLAS)
                except Exception as e:
                    errores.append(f"No se pudo abrir la plantilla para '{concepto}': {e}")
                    continue

                hojas_plantilla        = list(wb_concepto.sheetnames)
                nombres_usados         = set(wb_concepto.sheetnames)
                hojas_generadas        = []
                periodicidades         = []
                tiempos                = []

                for equipo_dict in lista_equipos:
                    activo_equipo = equipo_dict.get("# ACTIVO", "SIN_ACTIVO")
                    analizadores_equipo = analizadores_por_concepto.get(concepto, [])
                    try:
                        ws_base   = wb_concepto[pestana]
                        ws_equipo = wb_concepto.copy_worksheet(ws_base)
                        _copiar_imagenes_hoja(ws_base, ws_equipo)
                        ws_equipo.title = _construir_nombre_hoja(equipo_dict, nombres_usados)
                        _llenar_hoja_verificacion(
                            ws_equipo, equipo_dict, ingeniero, jefe, hospital,
                            nombre_plantilla=pestana, analizadores=analizadores_equipo
                        )
                        hojas_generadas.append(ws_equipo.title)
                        exitos += 1

                        if p := str(equipo_dict.get("PERIODICIDAD", "")).strip():
                            periodicidades.append(p)
                        if t := str(equipo_dict.get("TIEMPO MANTENIMIENTO", "")).strip():
                            tiempos.append(t)
                    except Exception as e:
                        errores.append(f"Error en {activo_equipo}: {e}")

                if hojas_generadas:
                    for hoja in hojas_plantilla:
                        if hoja in wb_concepto.sheetnames:
                            del wb_concepto[hoja]

                    periodicidades_unicas = list(dict.fromkeys(periodicidades))
                    tiempos_unicos        = list(dict.fromkeys(tiempos))
                    zip_file.writestr(
                        f"{carpeta_equipo}/datos_mantenimiento.txt",
                        "\n".join([
                            f"Concepto: {concepto}",
                            f"Periodicidad: {', '.join(periodicidades_unicas) or 'No definida'}",
                            f"Tiempo de mantenimiento: {', '.join(tiempos_unicos) or 'No capturado'}",
                        ])
                    )

                    nombre_excel = f"{tipo_equipo}.xlsx"
                    os.makedirs(RUTA_REPORTES, exist_ok=True)
                    ruta_concepto = os.path.join(RUTA_REPORTES, nombre_excel)
                    sufijo = 2
                    while os.path.exists(ruta_concepto):
                        nombre_excel  = f"{tipo_equipo}_{sufijo}.xlsx"
                        ruta_concepto = os.path.join(RUTA_REPORTES, nombre_excel)
                        sufijo += 1

                    wb_concepto.save(ruta_concepto)
                    zip_file.write(ruta_concepto, arcname=f"{carpeta_equipo}/{nombre_excel}")
                    if os.path.exists(ruta_concepto):
                        os.remove(ruta_concepto)

                try:
                    wb_concepto.close()
                except Exception:
                    pass

        if hacer_etiquetas:
            status_text.text("Generando etiquetas PDF...")
            ruta_pdf = crear_etiquetas_pdf(equipos, ingeniero,
                                           fecha_mantenimiento_base=fecha_mantenimiento_base)
            if os.path.exists(ruta_pdf):
                zip_file.write(ruta_pdf, arcname=f"{nombre_carpeta}/etiquetas_mantenimiento.pdf")
                os.remove(ruta_pdf)

        status_text.empty()

    buffer_zip.seek(0)
    return buffer_zip, errores, exitos


# ── Etiquetas PDF ─────────────────────────────────────────────────────────────

def _ajustar_texto_a_ancho(canvas_obj, texto, fuente, tamano, ancho_maximo, sufijo="..."):
    texto_base = str(texto or "")
    if ancho_maximo <= 0:
        return ""
    if canvas_obj.stringWidth(texto_base, fuente, tamano) <= ancho_maximo:
        return texto_base

    texto_base   = texto_base.strip()
    ancho_sufijo = canvas_obj.stringWidth(sufijo, fuente, tamano)
    if ancho_sufijo >= ancho_maximo:
        return ""

    fin = len(texto_base)
    while fin > 0:
        candidato = texto_base[:fin].rstrip() + sufijo
        if canvas_obj.stringWidth(candidato, fuente, tamano) <= ancho_maximo:
            return candidato
        fin -= 1
    return ""


def crear_etiquetas_pdf(equipos, ingeniero=None, fecha_mantenimiento_base=None):
    RUTA_PDF = os.path.join(RUTA_REPORTES, "etiquetas_mantenimiento.pdf")
    os.makedirs(RUTA_REPORTES, exist_ok=True)

    ancho_pagina, alto_pagina = letter

    ancho_etiqueta = 69 * mm
    alto_etiqueta  = 43 * mm
    columnas       = 3
    filas          = 6
    margen_x       = 2 * mm
    margen_y       = 0.5 * mm
    espacio_col    = 2 * mm
    espacio_fila   = 2 * mm
    alto_header    = 8 * mm
    alto_pie       = 4 * mm
    fuente_campos  = 6
    fuente_pie     = 5
    fuente_contacto = 4
    espacio_campos = 3 * mm

    c           = pdf_canvas.Canvas(RUTA_PDF, pagesize=letter)
    col_actual  = 0
    fila_actual = 0
    fecha_hoy   = date.today()

    for _, row in equipos.iterrows():
        x  = margen_x + col_actual * (ancho_etiqueta + espacio_col)
        y  = alto_pagina - margen_y - alto_etiqueta - fila_actual * (alto_etiqueta + espacio_fila)
        aw = ancho_etiqueta
        ah = alto_etiqueta

        c.setLineWidth(0.5)
        c.setStrokeColorRGB(0.1, 0.1, 0.1)
        c.rect(x, y, aw, ah)

        ap = alto_pie
        c.setFillColorRGB(0.192, 0.509, 0.580)
        c.rect(x, y, aw, ap, fill=True, stroke=False)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica", fuente_pie)
        c.drawCentredString(x + aw/2, y + ap * 0.62,
                            "En caso de mal funcionamiento reporte esta unidad al")
        c.drawCentredString(x + aw/2, y + ap * 0.22,
                            "Departamento de Ingenieria Biomedica de su Hospital")

        # ── Sección fecha / próximo ──────────────────────────────────────────
        y_fecha      = y + 1 * mm
        periodicidad = str(row.get("PERIODICIDAD", "Anual")).strip()
        fecha_etiqueta  = resolver_fecha_base_por_equipo(row, fecha_hoy, fecha_mantenimiento_base)
        fecha_siguiente = calcular_siguiente_mantenimiento(fecha_etiqueta, periodicidad)

        c.setFillColorRGB(0, 0, 0)

        # Fecha — etiqueta centrada arriba, línea abajo, valor sobre la línea
        c.setFont("Helvetica", fuente_campos - 1)
        c.drawCentredString(x + 17*mm, y_fecha + 8*mm, "Fecha Mantenimineto:")
        c.setFont("Helvetica", fuente_campos)
        c.drawCentredString(x + 17*mm, y_fecha + 4.5*mm, formato_mes_anio(fecha_etiqueta))
        c.line(x + 2*mm, y_fecha + 3.5*mm, x + 32*mm, y_fecha + 3.5*mm)

        # Próximo — etiqueta centrada arriba, línea abajo, valor sobre la línea
        c.setFont("Helvetica", fuente_campos - 1)
        c.drawCentredString(x + aw - 17*mm, y_fecha + 8*mm, "Próximo Mantenimiento:")
        c.setFont("Helvetica", fuente_campos)
        c.drawCentredString(x + aw - 17*mm, y_fecha + 4.5*mm, formato_mes_anio(fecha_siguiente))
        c.line(x + 35*mm, y_fecha + 3.5*mm, x + aw - 2*mm, y_fecha + 3.5*mm)

        ah_h   = alto_header
        y_head = y + ah - ah_h
        c.setFillColorRGB(1, 1, 1)
        c.rect(x + 0.5, y_head, aw - 1, ah_h - 0.5, fill=True, stroke=False)
        c.setLineWidth(0.5)
        c.setStrokeColorRGB(0.1, 0.1, 0.1)
        c.rect(x, y, aw, ah)

        try:
            logo = ImageReader("image_5976e1.png")
            c.drawImage(logo, x + 1*mm, y_head + 1.5*mm,
                        width=22*mm, height=ah_h - 3*mm,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica", fuente_contacto)
        c.drawRightString(x + aw - 1*mm, y_head + ah_h * 0.75, "Contacto:")
        c.drawRightString(x + aw - 1*mm, y_head + ah_h * 0.45, "+52(33) 38328790")
        c.drawRightString(x + aw - 1*mm, y_head + ah_h * 0.15, "018000051015")

        campos = [
            ("Equipo",         str(row.get("CONCEPTO", ""))),
            ("Marca",          str(row.get("MARCA", ""))),
            ("Modelo",         str(row.get("MODELO", ""))),
            ("No. Serie",      str(row.get("No. DE SERIE", ""))),
            ("No. Inventario", str(row.get("# ACTIVO", ""))),
            ("Area",           _construir_ubicacion_sububicacion(row)),
        ]

        field_step = 3 * mm
        y_campos   = y_head - 2 * mm
        for i, (etiqueta, valor) in enumerate(campos):
            yc = y_campos if i == 0 else y_campos - i * field_step
            ancho_label = c.stringWidth(f"{etiqueta}:", "Helvetica-Bold", fuente_campos)
            c.setFont("Helvetica-Bold", fuente_campos)
            c.setFillColorRGB(0, 0, 0)
            c.drawString(x + 2*mm, yc, f"{etiqueta}:")
            lx = x + 2*mm + ancho_label + 1*mm
            c.line(lx, yc - 0.8*mm, x + aw - 2*mm, yc - 0.8*mm)
            c.setFont("Helvetica", fuente_campos)
            valor_ajustado = _ajustar_texto_a_ancho(
                c, valor, "Helvetica", fuente_campos, (x + aw - 2*mm) - (lx + 1*mm)
            )
            c.drawString(lx + 1*mm, yc, valor_ajustado)

        y_last_field = y_campos - (len(campos) - 1) * field_step
        y_title = y_last_field - 2*mm - field_step + (field_step / 2)
        y_name  = y_title - 3*mm
        y_line  = y_name  - 0.5*mm
        c.setFont("Helvetica", fuente_campos)
        c.drawCentredString(x + aw/2, y_title, "Mantenimiento Preventivo realizado por:")
        c.setFont("Helvetica-Bold", fuente_campos)
        c.drawCentredString(x + aw/2, y_name, ingeniero or "")
        c.setLineWidth(0.5)
        c.line(x + 3*mm, y_line, x + aw - 3*mm, y_line)

        col_actual += 1
        if col_actual >= columnas:
            col_actual  = 0
            fila_actual += 1
        if fila_actual >= filas:
            fila_actual = 0
            col_actual  = 0
            c.showPage()

    c.save()
    return RUTA_PDF